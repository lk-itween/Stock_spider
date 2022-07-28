import os
import time
import json
import re
import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import requests
from faker import Faker
from multiprocessing.dummy import Pool as ThreadPool

import warnings

warnings.filterwarnings("ignore")


class ebscn_db(object):
    __slots__ = ['page_end', 'page_start', 'df_data', 'file_path', 'label2show']
    np.set_printoptions(suppress=True, threshold=1000)
    pd.set_option('display.max_columns', 10000, 'display.max_rows', 10000, 'display.float_format', lambda x: '%.2f' % x)

    date0 = datetime.datetime.now()
    date = date0.strftime('%Y%m%d')
    url = 'http://www.ebscn.com/servlet/json?'
    page_size = 5

    def __init__(self, page_start: str, page_end: str, path: str):
        self.page_end = page_end
        self.page_start = page_start
        self.file_path = path
        self.label2show = '开始抓取信息...\n'

    # 多线程，设置线程池
    def thread_main(self):
        if not os.path.exists(f'{self.file_path}/光大证券{self.date}.xlsx'):
            pd.DataFrame().to_excel(f'{self.file_path}/光大证券{self.date}.xlsx',sheet_name='可担保物')
        p_data = list(self.post_data(N1=4196))[0]
        html = self.get_html_ajex(self.url, p_data).decode('utf-8')
        data, page_total = self.parse_html(html)
#         print(page_total)
#         page_total = 4196
#         self.df_data = self.running_main(self.post_data(N1=page_total))
        self.df_data = data
        self.df_data = self.modify_field(self.df_data)
        self.data_save(self.df_data, self.file_path, label='可担保物')
        self.label2show = '所有信息抓取完毕。'
        return self.df_data

    # 运行函数
    def running_main(self, p_data):
        time.sleep(np.random.random() * 1.1 + 1)
#         self.label2show = f"正在抓取第{p_data}页...\n"
#         print(self.label2show)
        try:
            html = self.get_html_ajex(self.url, p_data).decode('utf-8')
        except AttributeError:
            time.sleep(15)
            return self.running_main(p_data)
        data, page_total = self.parse_html(html)
#         self.label2show = f"第{p_data}页已抓取。\n"
        time.sleep(np.random.random() + 2.5)
        return data

    # 构造请求数据字典
    def post_data(self, page_num=1, N1=20):
        data = f'funcNo=501003&curtPageNo={page_num}&numPerPage={N1}'
        yield data

    # 请求初始网页
    def get_html_ajex(self, url, data=None):
        UA = Faker(locale='zh_CN').chrome(version_from=80, version_to=84, build_from=3600, build_to=4200)
        try:
            headers = {
                'Host': 'www.ebscn.com',
                'Connection': 'keep-alive',
#                 'Content-Length': '62',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
#                 'X-Requested-With': 'XMLHttpRequest',
#                 'Origin': 'http://www.ebscn.com',
#                 'Referer': 'http://www.ebscn.com/main/ourBusiness/xyyw/rzrq/cyxx/index.html',
                'Accept-Encoding': 'gzip, deflate',
                'Accept-Language': 'zh-CN,zh;q=0.9',
#                 'Upgrade-Insecure-Requests': '1',
                'Cache-Control': 'max-age=0',
#                 'Cookie':'ekV2Di3wYMf3S=5GDX29aP6TmqI321WtDDjLvFJEbdeljKN5WLOjy0fvocsX1X5g2S1ufV1O.FqMcXyUbrd4f.o3NHrAjTQysEUca; Hm_lvt_7f2befad9fea9875c861e4f22a57a7e5=1607996225; Hm_lpvt_7f2befad9fea9875c861e4f22a57a7e5=1608010653; ekV2Di3wYMf3T=5Uefhl2d31yaqqqmCRAfF.GxAp386dCwJkF.dOdpIs6aIOIUoRRKutD7NnqzEqgIkC3bkWkcM.__oyzzl8oP2Bsa4BPFR58wQjwYC.ZaSrLj6vKLGIHFcDDuK5Jp34fhlNTICtbsCaKEHHwERfLCP2UUmgGdMAc3DooPmkLuu8pudC2tqYSUksw9k8mpynSarQH7Nu30ptEMRb6uBBIKapqdfG6ZCf4Akz.gEXXd3G.d5KuKeXRNk6V84063MyIPKtYt9sPgVe8J7ojQ0Xs9frafqry6D176UB0lma2p6.2jAyDUvVuObdzWQwNLsZLUFW',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4021.2 Safari/537.36',
            }
            Cookies = input('请输入Cookie:').split('Cookie: ')[1]
            headers['Cookie'] = Cookies
            time.sleep(np.random.random() * 1.2 + 0.5)
            res = requests.get(url+data, headers=headers)
            sta_code = res.status_code
            print(url)
            print(data)
            print(headers)
            print(sta_code)
            if sta_code == 405:
                url = url.replace('http:', 'https:')
                res = requests.get(url+data, headers=headers)
                sta_code = res.status_code
            if sta_code == 200:
                return res.content
            return None
        except:
            return None

    # 解析初始网页 json 数据
    def parse_html(self, html):
        df_json = json.loads(html)['results'][0]
        page_total = df_json['totalRows'] if type(df_json['totalRows']) == int else int(df_json['totalRows'].strip(',').strip(' '))
        df = pd.DataFrame(df_json['data'])
#         display(df)
        df_columns = {'sec_market':'市场','sec_code': '证券代码', 'sec_short_name': '证券简称','conversion_rate': '折算率','trade_date': '日期'}
        df.rename(columns=df_columns, inplace=True)
        df = df[['市场', '证券代码', '证券简称', '折算率', '日期']]
        df.replace(r'^$',np.nan,regex=True,inplace=True)
        df.dropna(inplace=True)
        df['市场'] = df['市场'].map(lambda x: '深圳' if x == '深A' else '上海')
#         df['证券简称'] = df['证券简称'].map(self.uni2utf)
#         df['折算率'] = df['折算率'].map(lambda x: str(round(float(x)*100,2))+'%')
        df.replace(' ','',regex=True,inplace=True)
        df.replace(['0.0%','0%','-'],'0',inplace=True)
        return df , page_total

    # unicode字符转换
    def uni2utf(self,df):
        str1 = ''.join('\\u' + i for i in re.findall('u([a-zA-Z0-9]{4})',df))
        return json.loads(f'"{str1}"')
    
    # 字段格式规格化
    def modify_field(self, df):
        df.replace(['（', '）'], ['(', ')'], regex=True, inplace=True)
        date_columns = [i for i in df.columns if '日期' in i]
        for i in date_columns:
            df[i] = df[i].str.replace('-', '')
        return df

    # 数据储存为 excel 格式
    def data_save(self, df, path, label='融资标的'):
        with pd.ExcelWriter(
                f'{path}/光大证券{self.date}.xlsx', engine='openpyxl', mode='w+') as writer:
            book = load_workbook(f'{path}/光大证券{self.date}.xlsx')
            writer.book = book
            wb = writer.book
            try:
                ws = wb[label]
                wb.remove(ws)
            except:
                pass
            df.to_excel(writer, index=False, sheet_name=label)


class ebscn_rzrq(object):
    __slots__ = ['page_end', 'page_start', 'df_data', 'file_path', 'label2show']
    np.set_printoptions(suppress=True, threshold=1000)
    pd.set_option('display.max_columns', 10000, 'display.max_rows', 10000, 'display.float_format', lambda x: '%.2f' % x)

    date0 = datetime.datetime.now()
    date = date0.strftime('%Y%m%d')
    date1 = date0.strftime('%Y-%m-%d')
    url = 'http://www.ebscn.com/servlet/json?'
    page_size = 5

    def __init__(self, page_start: str, page_end: str, path: str):
        self.page_end = page_end
        self.page_start = page_start
        self.file_path = path
        self.label2show = '开始抓取信息...\n'

    # 多线程，设置线程池
    def thread_main(self):
        if not os.path.exists(f'{self.file_path}/光大证券{self.date}.xlsx'):
            pd.DataFrame().to_excel(f'{self.file_path}/光大证券{self.date}.xlsx',sheet_name='融资标的')
        p_data = list(self.post_data(N1=1981))[0]
        html = self.get_html_ajex(self.url, p_data).decode('utf-8')
#         return html
        data, page_total = self.parse_html(html)
#         print(page_total)
#         page_total = 4196
#         self.df_data = self.running_main(self.post_data(N1=page_total))
        self.df_data = data
        df_rz, df_rq = self.modify_field(self.df_data)
        self.data_save(df_rz, self.file_path, label='融资标的')
        self.data_save(df_rq, self.file_path, label='融券标的')
        self.label2show = '所有信息抓取完毕。'

    # 运行函数
    def running_main(self, data_range):
        url = self.url_rz if data_range[1] == 'rzb' else self.url_rq
        p_data = data_range[0]
        time.sleep(np.random.random() * 1.1 + 1)
        self.label2show = f"正在抓取第{p_data}页...\n"
        try:
            html = self.get_html_ajex(url, data_range[1], p_data).decode('utf-8')
        except AttributeError:
            time.sleep(15)
            return self.running_main(p_data)
        data, page_total = self.parse_html(html)
        self.label2show = f"第{p_data}页已抓取。\n"
        time.sleep(np.random.random() + 2.5)
        return data

    # 构造请求数据字典
    def post_data(self, page_num=1, N1=20):
        data = f'funcNo=501002&curtPageNo={page_num}&numPerPage={N1}'
        yield data

    # 请求初始网页
    def get_html_ajex(self, url, data=None):
        UA = Faker(locale='zh_CN').chrome(version_from=80, version_to=84, build_from=3600, build_to=4200)
        try:
            headers = {
                'Host': 'www.ebscn.com',
                'Connection': 'keep-alive',
#                 'Content-Length': '62',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
#                 'X-Requested-With': 'XMLHttpRequest',
#                 'Origin': 'http://www.ebscn.com',
#                 'Referer': 'http://www.ebscn.com/main/ourBusiness/xyyw/rzrq/cyxx/index.html',
                'Accept-Encoding': 'gzip, deflate',
                'Accept-Language': 'zh-CN,zh;q=0.9',
#                 'Upgrade-Insecure-Requests': '1',
                'Cache-Control': 'max-age=0',
#                 'Cookie':'ekV2Di3wYMf3S=5GDX29aP6TmqI321WtDDjLvFJEbdeljKN5WLOjy0fvocsX1X5g2S1ufV1O.FqMcXyUbrd4f.o3NHrAjTQysEUca; Hm_lvt_7f2befad9fea9875c861e4f22a57a7e5=1607996225; Hm_lpvt_7f2befad9fea9875c861e4f22a57a7e5=1608010653; ekV2Di3wYMf3T=5Uefhl2d31yaqqqmCRAfF.GxAp386dCwJkF.dOdpIs6aIOIUoRRKutD7NnqzEqgIkC3bkWkcM.__oyzzl8oP2Bsa4BPFR58wQjwYC.ZaSrLj6vKLGIHFcDDuK5Jp34fhlNTICtbsCaKEHHwERfLCP2UUmgGdMAc3DooPmkLuu8pudC2tqYSUksw9k8mpynSarQH7Nu30ptEMRb6uBBIKapqdfG6ZCf4Akz.gEXXd3G.d5KuKeXRNk6V84063MyIPKtYt9sPgVe8J7ojQ0Xs9frafqry6D176UB0lma2p6.2jAyDUvVuObdzWQwNLsZLUFW',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4021.2 Safari/537.36',
            }
            Cookies = input('请输入Cookie:').split('Cookie: ')[1]
            headers['Cookie'] = Cookies
            time.sleep(np.random.random() * 1.2 + 0.5)
            res = requests.get(url+data, headers=headers)
            sta_code = res.status_code
            print(url)
            print(data)
            print(headers)
            print(sta_code)
            if sta_code == 405:
                url = url.replace('http:', 'https:')
                res = requests.get(url+data, headers=headers)
                sta_code = res.status_code
            if sta_code == 200:
                return res.content
            return None
        except:
            return None

    # 解析初始网页 json 数据
    def parse_html(self, html):
        df_json = json.loads(html)['results'][0]
        page_total = df_json['totalRows'] if type(df_json['totalRows']) == int else int(df_json['totalRows'].strip(',').strip(' '))
        df = pd.DataFrame(df_json['data'])
        df_columns = {'sec_market':'市场','sec_code': '证券代码', 'sec_abbreviation': '证券简称',
                      'financing_sec':'融资标的','marriage_sec': '融券标的','sec_date': '日期'}
        df.rename(columns=df_columns, inplace=True)
        df = df[['市场', '证券代码', '证券简称', '融资标的','融券标的', '日期']]
        df.replace(r'^$',np.nan,regex=True,inplace=True)
        df.dropna(inplace=True)
        df['市场'] = df['市场'].map(lambda x: '深圳' if x == '深A' else '上海')
        df.replace(' ','',regex=True,inplace=True)
        df.replace(['0.0%','0%','-'],'0',inplace=True)
        return df, page_total

    # 字段格式规格化
    def modify_field(self, df, label='融资'):
        df.replace(['（', '）'], ['(', ')'], regex=True, inplace=True)
        date_columns = [i for i in df.columns if '日期' in i]
        for i in date_columns:
            df[i] = df[i].str.replace('-', '')
        df_rz = df[['市场', '证券代码', '证券简称', '融资标的', '日期']]
        df_rq = df[['市场', '证券代码', '证券简称', '融券标的', '日期']]
        return df_rz,df_rq

    # 数据储存为 excel 格式
    def data_save(self, df, path, label='融资标的'):
        with pd.ExcelWriter(
                f'{path}/光大证券{self.date}.xlsx', engine='openpyxl', mode='w+') as writer:
            book = load_workbook(f'{path}/光大证券{self.date}.xlsx')
            writer.book = book
            wb = writer.book
            try:
                ws = wb[label]
                wb.remove(ws)
            except:
                pass
            df.to_excel(writer, index=False, sheet_name=label)


# if __name__ == '__main__':
#     run = cczq_db('1', '2', '.')
#     run1 = cczq_rzrq('1', '2', '.')
#     run.thread_main()
#     run1.thread_main()